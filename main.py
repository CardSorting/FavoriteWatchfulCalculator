import os
import requests
import openai
import boto3
import json
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image
import pytesseract
import datetime
import cv2
import logging
from typing import Tuple

# Set up logging
logging.basicConfig(level=logging.INFO)

# Set up API keys
openai.api_key = os.getenv('OPENAI_API_KEY')
countdown_api_key = os.getenv('COUNTDOWN_API_KEY')
backblaze_app_id = os.getenv('BACKBLAZE_APP_ID')
backblaze_app_key = os.getenv('BACKBLAZE_APP_KEY')

# Initialize Backblaze B2 client
b2_client = boto3.client(
    's3',
    endpoint_url='https://s3.us-east-005.backblazeb2.com',
    aws_access_key_id=backblaze_app_id,
    aws_secret_access_key=backblaze_app_key
)

bucket_name = 'your-bucket-name'  # Replace with your Backblaze B2 bucket name
filename = 'yu_gi_oh_cards.xlsx'  # Extracted filename to a variable

# Function to upload image to Backblaze B2
def upload_image_to_backblaze(image_path: str) -> str:
    try:
        image_name = f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{os.path.basename(image_path)}"
        b2_client.upload_file(image_path, bucket_name, image_name)
        return f"https://s3.us-east-005.backblazeb2.com/{bucket_name}/{image_name}"
    except Exception as e:
        logging.error(f"Failed to upload image to Backblaze: {e}")
        return ""

# Function to capture an image from the webcam
def capture_image_from_webcam() -> str:
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        logging.error("Cannot open camera")
        return ""
    ret, frame = cap.read()
    if not ret:
        logging.error("Can't receive frame (stream end?). Exiting ...")
        return ""
    image_path = "captured_card.jpg"
    cv2.imwrite(image_path, frame)
    cap.release()
    cv2.destroyAllWindows()
    return image_path

# Function to perform OCR on an image of the card
def scan_card_image(image_path: str) -> str:
    try:
        image = Image.open(image_path)
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        logging.error(f"Failed to scan card image: {e}")
        return ""

# Function to extract card details using OpenAI API
def extract_card_details(text: str) -> dict:
    try:
        response = openai.Completion.create(
            model="text-davinci-003",
            prompt=f"Extract the Yu-Gi-Oh! card name and serial number from the following text:\n{text}",
            max_tokens=50
        )
        details = response['choices'][0]['text'].strip().split(',')
        if len(details) != 2:
            return {"name": "Unknown", "serial_number": "Unknown"}
        card_name, serial_number = details
        return {
            "name": card_name.strip(),
            "serial_number": serial_number.strip()
        }
    except Exception as e:
        logging.error(f"Failed to extract card details: {e}")
        return {"name": "Unknown", "serial_number": "Unknown"}

# Function to retrieve card prices from eBay using Countdown API
def get_card_price(card_name: str) -> float:
    try:
        params = {
            'api_key': countdown_api_key,
            'type': 'search',
            'ebay_domain': 'ebay.com',
            'search_term': card_name
        }
        api_result = requests.get('https://api.countdownapi.com/request', params=params)
        result_json = api_result.json()

        # Extract prices from search results, limit to first 10 results for efficiency
        prices = [item['price'] for item in result_json.get('search_results', [])[:10] if 'price' in item]

        if prices:
            average_price = sum(prices) / len(prices)
            return round(average_price, 2)
        else:
            return 0.0
    except Exception as e:
        logging.error(f"Failed to get card price: {e}")
        return 0.0

# Function to create or load the workbook
def create_or_load_workbook(filename: str) -> Tuple[Workbook, Worksheet]:
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        if not isinstance(sheet, Worksheet):
            raise ValueError("Loaded sheet is not a Worksheet")
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        headers = [
            "*Action(SiteID=US|Country=US|Currency=USD|Version=941)", "CustomLabel", "*Category", "StoreCategory",
            "*Title", "Subtitle", "Relationship", "RelationshipDetails", "*ConditionID", "Condition Descriptor Name 1",
            "Condition Descriptor Value 1", "CD:Professional Grader - (ID: 27501)", "CD:Grade - (ID: 27502)",
            "CDA:Certification Number - (ID: 27503)", "CD:Card Condition - (ID: 40001)", "*C:Franchise", "C:Set",
            "C:Manufacturer", "C:Year Manufactured", "C:Character", "C:TV Show", "C:Autograph Authentication",
            "C:Grade", "C:Features", "C:Parallel/Variety", "C:Featured Person/Artist", "C:Autographed", "C:Type",
            "C:Card Number", "C:Card Name", "C:Movie", "C:Age Level", "C:Signed By", "C:Material", "C:Genre",
            "C:Graded", "C:Card Size", "C:Language", "C:Manufacturered in", "P:UPC", "Start Price", "Quantity",
            "Item photo URL", "P:EAN", "Shipping Profile Name", "Return Profile Name", "Payment Profile Name",
            "ShippingType", "ShippingService", "ShippingServiceCost", "ShippingServiceAdditionalCost",
            "ShippingServicePriority", "Max Dispatch Time", "Returns Accepted Option", "Returns Within Option",
            "Refund Option", "Return Shipping Cost Paid By", "ListingDuration", "Location", "Description"
        ]
        sheet.append(headers)
    return workbook, sheet

# Function to add card details to the workbook
def add_card_to_workbook(workbook: Workbook, sheet: Worksheet, card_details: dict, price: float, image_url: str):
    try:
        row = [
            "Add", "", "183454", "", card_details["name"], "", "", "", "4000", "40001", "400010", "", "", "", "400010",
            "Yu-Gi-Oh!", "", "Konami", "2002", "", "Yu-Gi-Oh!", "", "", "", "", "No", "Trading Card",
            card_details["serial_number"], card_details["name"], "", "10+", "", "Card Stock", "Collectible Card Game", "No",
            "Standard", "English", "Japan", "", price, "1", image_url, "", "Shipping-Default", "Return-Default", "Payment-Policy-Default",
            "Flat", "USPSFirstClass", "3.50", "0.50", "1", "1", "ReturnsAccepted", "Days_30", "MoneyBack", "Buyer",
            "Days_7", "", f"Card Name: {card_details['name']}, Serial Number: {card_details['serial_number']}"
        ]
        sheet.append(row)
    except Exception as e:
        logging.error(f"Failed to add card to workbook: {e}")

# Main function to coordinate the process
def main():
    workbook, sheet = create_or_load_workbook(filename)

    # Capture image from webcam
    image_path = capture_image_from_webcam()

    if not image_path:
        logging.error("Image capture failed. Exiting.")
        return

    image_url = upload_image_to_backblaze(image_path)
    if not image_url:
        logging.error("Image upload failed. Exiting.")
        return

    text = scan_card_image(image_path)
    if not text:
        logging.error("Image scanning failed. Exiting.")
        return

    card_details = extract_card_details(text)
    price = get_card_price(card_details["name"])
    if price:
        add_card_to_workbook(workbook, sheet, card_details, price, image_url)
        workbook.save(filename)
        logging.info(f"{card_details['name']}: ${price}. Scan another card or compile the list?")
    else:
        logging.error(f"Price for {card_details['name']} not found. Please try another card or check the card details.")

if __name__ == "__main__":
    main()